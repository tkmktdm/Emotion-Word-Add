import openpyxl
import pprint
import os
import glob
import numpy as np
import itertools

class XlsxExist:

    def xlsxCheck(self):
        xlsxfile = os.path.exists('words_emo.xlsx')
        print(os.path.abspath('words_emo.xlsx'))
        if xlsxfile:
            print('ok')
        else:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = '感情分析'
            wb.save('words_emo.xlsx')
            print('newfile create')

class XlsxWR:
    def xlsxRead(self):
        wb = openpyxl.load_workbook('words_emo.xlsx')
        sheet = wb['感情分析']
        def get_value_list(t_2d):
            return([[cell.value for cell in row] for row in t_2d])

        def get_list_2d(sheet, start_row, end_row, start_col, end_col):
            return get_value_list(sheet.iter_rows(min_row=start_row,
                                                max_row=end_row,
                                                min_col=start_col,
                                                max_col=end_col))
        w_dic = get_list_2d(sheet, 0, 0, 1, 3)
        return w_dic

    def xlsxWrite(self, d_2d):
        wb = openpyxl.load_workbook('words_emo.xlsx')
        sheet = wb['感情分析']
        def write_list_2d(sheet, d_2d, start_row, start_col):
            for y, row in enumerate(d_2d):
                for x, cell in enumerate(row):
                    sheet.cell(row=start_row + y,
                            column=start_col + x,
                            value=d_2d[y][x])
        write_list_2d(sheet, d_2d, 1, 1)
        wb.save('words_emo.xlsx')
        print('save')

'''xe = XlsxExist()
xe.xlsxCheck()
xwr = XlsxWR()
worddic = xwr.xlsxRead()
#入力された文章の処理が完了したら以下の処理を実行すると保存できる。
words = [['鬼', '名詞', 0], ['滅', '名詞', 0], ['刃', '名詞', 0], ['最高', '名詞', 0]]
for w in words:
    worddic.append(w)

xwr.xlsxWrite(worddic)'''
