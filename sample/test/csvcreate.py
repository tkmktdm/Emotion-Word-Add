import openpyxl
import pprint
import os
import glob
import numpy as np

#data = np.loadtxt('input.csv', delimiter=',', skiprow=0)



word = [['のど', '名詞', 0], ['コーラ', '名詞', 0], ['コロナ', '名詞', -1]]

class XlsxExist:

    def xlsxCheck(self):
        xlsxfile = os.path.exists('words_emo.xlsx')
        if xlsxfile:
            #wb = openpyxl.load_workbook('words_emo.xlsx')
            #sheet = wb['感情分析']
            #if wb.sheetnames:
            #    l_2d1 = get_list_2d(sheet, 2, 50, 1, 2)

            #test = xlsxRead()
            #return test
            #xlsxWrite()   

            #else:
            #    write_list_2d(sheet, word, 1, 1)
            #    wb.save('words_emo.xlsx')
                #print('ng')
            #sheet = wb['感情分類']
            print('ok')
        else:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = '感情分析'
            wb.save('words_emo.xlsx')
            print('newfile create')
    '''def xlsxRead(self):
        wb = openpyxl.load_workbook('words_emo.xlsx')
        sheet = wb['感情分析']
        def get_value_list(t_2d):
            return([[cell.value for cell in row] for row in t_2d])

        def get_list_2d(sheet, start_row, end_row, start_col, end_col):
            return get_value_list(sheet.iter_rows(min_row=start_row,
                                                max_row=end_row,
                                                min_col=start_col,
                                                max_col=end_col))
        l_2d1 = get_list_2d(sheet, 2, 50, 1, 2)
        return l_2d1

    def xlsxWrite(self):
        wb = openpyxl.load_workbook('words_emo.xlsx')
        sheet = wb['感情分析']
        def write_list_2d(sheet, d_2d, start_row, start_col):
            for y, row in enumerate(d_2d):
                for x, cell in enumerate(row):
                    sheet.cell(row=start_row + y,
                            column=start_col + x,
                            value=d_2d[y][x])'''

class XlsxWR:
    def xlsxRead(self):
        wb = openpyxl.load_workbook('words_emo.xlsx')
        sheet = wb['感情分析']
        def get_value_list(t_2d):
            return([[cell.value for cell in row] for row in t_2d])

        def get_list_2d(sheet, start_row, end_row, start_col, end_col):
            return get_value_list(sheet.iter_rows(min_row=start_row,
                                                max_row=end_row,
                                                min_col=start_col,
                                                max_col=end_col))
        w_dic = get_list_2d(sheet, 0, 0, 1, 3)
        return w_dic

    def xlsxWrite(self, d_2d):
        wb = openpyxl.load_workbook('words_emo.xlsx')
        sheet = wb['感情分析']
        def write_list_2d(sheet, d_2d, start_row, start_col):
            for y, row in enumerate(d_2d):
                for x, cell in enumerate(row):
                    sheet.cell(row=start_row + y,
                            column=start_col + x,
                            value=d_2d[y][x])
        print(d_2d)
        write_list_2d(sheet, d_2d, 1, 1)
        wb.save('words_emo.xlsx')
        print('save')

'''    def get_value_list(t_2d):
        return([[cell.value for cell in row] for row in t_2d])

    def get_list_2d(sheet, start_row, end_row, start_col, end_col):
        return get_value_list(sheet.iter_rows(min_row=start_row,
                                            max_row=end_row,
                                            min_col=start_col,
                                            max_col=end_col))
    def write_list_2d(sheet, d_2d, start_row, start_col):
        for y, row in enumerate(d_2d):
            for x, cell in enumerate(row):
                sheet.cell(row=start_row + y,
                        column=start_col + x,
                        value=d_2d[y][x])'''

xe = XlsxExist()
xe.xlsxCheck()
xwr = XlsxWR()
worddic = xwr.xlsxRead()
t = np.array([['鬼', '名詞', 0], ['滅', '名詞', 0], ['刃', '名詞', 0], ['最高', '名詞', 0]])
#t = ['昔', '名詞', 0]
#t = [['鬼', '名詞', 0], ['滅', '名詞', 0], ['刃', '名詞', 0], ['最高', '名詞', 0]]
np.savetxt('test.csv', t)

#worddic.append(t)
#xwr.xlsxWrite(worddic)

'''
    def get_value_list(t_2d):
        return([[cell.value for cell in row] for row in t_2d])

    def get_list_2d(sheet, start_row, end_row, start_col, end_col):
        return get_value_list(sheet.iter_rows(min_row=start_row,
                                            max_row=end_row,
                                            min_col=start_col,
                                            max_col=end_col))
''' 
    
'''
wb = openpyxl.load_workbook('words_emo.xlsx')
sheet = wb['感情分類']

def write_list_2d(sheet, d_2d, start_row, start_col):
    for y, row in enumerate(d_2d):
        for x, cell in enumerate(row):
            sheet.cell(row=start_row + y,
                       column=start_col + x,
                       value=d_2d[y][x])

write_list_2d(sheet, word, 1, 1)
#pprint.pprint(list(sheet.values), width=40)
wb.save('words_emo.xlsx')


#else:
#    print('false')

print('end')
#wb = openpyxl.load_workbook('words_emotion.xlsx')
#print(wb)
#wb.save('words_emotion.xlsx')
'''